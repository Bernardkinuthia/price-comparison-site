import csv, json, os, datetime, time
import hashlib, hmac, base64
from urllib.parse import quote, urlencode
import requests

# Try to import openpyxl for Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("📝 Note: Install openpyxl for Excel support (pip install openpyxl)")

# Amazon credentials from GitHub secrets
access_key = os.getenv("AMAZON_ACCESS_KEY")
secret_key = os.getenv("AMAZON_SECRET_KEY")
associate_tag = os.getenv("AMAZON_ASSOCIATE_TAG")

class AmazonPAAPI:
    def __init__(self, access_key, secret_key, associate_tag, region="us-east-1"):
        self.access_key = access_key
        self.secret_key = secret_key
        self.associate_tag = associate_tag
        self.region = region
        self.host = "webservices.amazon.com"
        self.marketplace = "www.amazon.com"
    
    def _get_signature(self, string_to_sign):
        """Generate AWS signature"""
        signing_key = ("AWS4" + self.secret_key).encode('utf-8')
        date_key = hmac.new(signing_key, datetime.datetime.utcnow().strftime('%Y%m%d').encode('utf-8'), hashlib.sha256).digest()
        region_key = hmac.new(date_key, self.region.encode('utf-8'), hashlib.sha256).digest()
        service_key = hmac.new(region_key, 'ProductAdvertisingAPI'.encode('utf-8'), hashlib.sha256).digest()
        request_key = hmac.new(service_key, 'aws4_request'.encode('utf-8'), hashlib.sha256).digest()
        signature = hmac.new(request_key, string_to_sign.encode('utf-8'), hashlib.sha256).hexdigest()
        return signature
    
    def get_items(self, asin_list):
        """Get item information from Amazon PA-API"""
        if isinstance(asin_list, str):
            asin_list = [asin_list]
        
        # Request payload
        payload = {
            "ItemIds": asin_list,
            "Resources": [
                "ItemInfo.Title",
                "Offers.Listings.Price"
            ],
            "PartnerTag": self.associate_tag,
            "PartnerType": "Associates",
            "Marketplace": self.marketplace
        }
        
        # Headers
        headers = {
            'Content-Type': 'application/json; charset=utf-8',
            'Content-Encoding': 'amz-1.0',
            'X-Amz-Target': 'com.amazon.paapi5.v1.ProductAdvertisingAPIv1.GetItems',
            'Host': self.host
        }
        
        # Current timestamp
        timestamp = datetime.datetime.utcnow()
        amz_date = timestamp.strftime('%Y%m%dT%H%M%SZ')
        date_stamp = timestamp.strftime('%Y%m%d')
        
        # Create canonical request
        canonical_headers = f"content-type:application/json; charset=utf-8\nhost:{self.host}\nx-amz-date:{amz_date}\nx-amz-target:com.amazon.paapi5.v1.ProductAdvertisingAPIv1.GetItems\n"
        signed_headers = "content-type;host;x-amz-date;x-amz-target"
        payload_json = json.dumps(payload, separators=(',', ':'))
        payload_hash = hashlib.sha256(payload_json.encode('utf-8')).hexdigest()
        
        canonical_request = f"POST\n/paapi5/getitems\n\n{canonical_headers}\n{signed_headers}\n{payload_hash}"
        
        # Create string to sign
        credential_scope = f"{date_stamp}/{self.region}/ProductAdvertisingAPI/aws4_request"
        string_to_sign = f"AWS4-HMAC-SHA256\n{amz_date}\n{credential_scope}\n{hashlib.sha256(canonical_request.encode('utf-8')).hexdigest()}"
        
        # Create signature
        signature = self._get_signature(string_to_sign)
        
        # Create authorization header
        authorization_header = f"AWS4-HMAC-SHA256 Credential={self.access_key}/{credential_scope}, SignedHeaders={signed_headers}, Signature={signature}"
        
        # Add auth headers
        headers['Authorization'] = authorization_header
        headers['X-Amz-Date'] = amz_date
        
        # Make request
        url = f"https://{self.host}/paapi5/getitems"
        
        try:
            response = requests.post(url, headers=headers, data=payload_json, timeout=30)
            if response.status_code == 200:
                return response.json()
            else:
                print(f"API Error {response.status_code}: {response.text}")
                return None
        except Exception as e:
            print(f"Request error: {e}")
            return None

def read_json_products(filename):
    """Read products from JSON file"""
    print(f"📖 Reading products from JSON file: {filename}")
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        products = []
        if isinstance(data, list):
            # Array of products
            for i, item in enumerate(data, 1):
                if isinstance(item, dict) and 'asin' in item:
                    asin = str(item['asin']).strip()
                    if asin and len(asin) == 10 and asin.isalnum():
                        products.append({
                            'asin': asin,
                            'title': item.get('title', ''),
                            'affiliate_url': item.get('affiliate_url', item.get('affiliate_link', ''))  # Support both names
                        })
                        print(f"  ✅ Product {i}: ASIN '{asin}' - Valid")
                    else:
                        print(f"  ⚠️ Product {i}: ASIN '{asin}' - Invalid")
        
        return products
    except Exception as e:
        print(f"❌ Error reading JSON: {e}")
        return []

def read_txt_products(filename):
    """Read products from plain text file"""
    print(f"📖 Reading products from text file: {filename}")
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        products = []
        for i, line in enumerate(lines, 1):
            line = line.strip()
            if not line or line.startswith('#'):  # Skip empty lines and comments
                continue
            
            # Support multiple formats:
            # 1. Just ASIN: B0C5HYBQMW
            # 2. ASIN|Title: B0C5HYBQMW|Product Name
            # 3. ASIN|Title|Link: B0C5HYBQMW|Product Name|https://link.com
            parts = line.split('|')
            
            asin = parts[0].strip()
            title = parts[1].strip() if len(parts) > 1 else ''
            affiliate_url = parts[2].strip() if len(parts) > 2 else ''
            
            if asin and len(asin) == 10 and asin.isalnum():
                products.append({
                    'asin': asin,
                    'title': title,
                    'affiliate_url': affiliate_url
                })
                print(f"  ✅ Line {i}: ASIN '{asin}' - Valid")
            else:
                print(f"  ⚠️ Line {i}: ASIN '{asin}' - Invalid")
        
        return products
    except Exception as e:
        print(f"❌ Error reading text file: {e}")
        return []

def read_excel_products(filename):
    """Read products from Excel file"""
    if not EXCEL_SUPPORT:
        print("❌ Excel support not available. Install openpyxl: pip install openpyxl")
        return []
    
    print(f"📖 Reading products from Excel file: {filename}")
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        
        # Find header row (usually row 1)
        headers = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                header_name = str(cell_value).strip().lower()
                if 'asin' in header_name:
                    headers['asin'] = col
                elif 'title' in header_name or 'name' in header_name:
                    headers['title'] = col
                elif 'link' in header_name or 'url' in header_name:
                    headers['affiliate_url'] = col
        
        print(f"🔍 Found headers: {headers}")
        
        if 'asin' not in headers:
            print("❌ No ASIN column found in Excel file")
            return []
        
        products = []
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            asin_cell = sheet.cell(row=row, column=headers['asin'])
            if not asin_cell.value:
                continue
            
            asin = str(asin_cell.value).strip()
            title = ''
            affiliate_url = ''
            
            if 'title' in headers:
                title_cell = sheet.cell(row=row, column=headers['title'])
                title = str(title_cell.value or '').strip()
            
            if 'affiliate_url' in headers:
                link_cell = sheet.cell(row=row, column=headers['affiliate_url'])
                affiliate_url = str(link_cell.value or '').strip()
            
            if asin and len(asin) == 10 and asin.isalnum():
                products.append({
                    'asin': asin,
                    'title': title,
                    'affiliate_url': affiliate_url
                })
                print(f"  ✅ Row {row}: ASIN '{asin}' - Valid")
            else:
                print(f"  ⚠️ Row {row}: ASIN '{asin}' - Invalid")
        
        return products
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return []

# Initialize API client
if not all([access_key, secret_key, associate_tag]):
    print("❌ Missing Amazon API credentials")
    exit(1)

amazon_api = AmazonPAAPI(access_key, secret_key, associate_tag)

# Try to read products from different file formats
products = []
possible_files = [
    ('products.json', read_json_products),
    ('products.txt', read_txt_products),
    ('products.xlsx', read_excel_products),
    ('products.xls', read_excel_products),
    ('products.csv', None)  # We'll handle CSV separately if needed
]

for filename, reader_func in possible_files:
    if os.path.exists(filename):
        print(f"\n🔍 Found file: {filename}")
        if reader_func:
            products = reader_func(filename)
            if products:
                print(f"✅ Successfully loaded {len(products)} products from {filename}")
                break
        elif filename.endswith('.csv'):
            # Fallback CSV reader with better encoding handling
            print("📖 Attempting to read CSV with multiple encodings...")
            for encoding in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
                try:
                    with open(filename, newline='', encoding=encoding) as csvfile:
                        reader = csv.DictReader(csvfile)
                        products = []
                        for row_num, row in enumerate(reader, 1):
                            # Try different possible column names for ASIN
                            asin_value = None
                            for possible_asin_key in ['asin', 'ASIN', 'uffeffasin']:
                                if possible_asin_key in row:
                                    asin_value = row[possible_asin_key]
                                    break
                            
                            if asin_value:
                                asin = str(asin_value).strip()
                                if len(asin) == 10 and asin.isalnum():
                                    products.append({
                                        'asin': asin,
                                        'title': row.get('title', ''),
                                        'affiliate_url': row.get('affiliate_url', row.get('affiliate_link', ''))  # Support both names
                                    })
                                    print(f"  ✅ Row {row_num}: ASIN '{asin}' - Valid")
                        
                        if products:
                            print(f"✅ Successfully read CSV with {encoding} encoding")
                            break
                except Exception as e:
                    print(f"  ⚠️ Failed with {encoding} encoding: {e}")
                    continue

if not products:
    print("\n❌ No valid products found in any supported file format!")
    print("\n📝 Supported formats and examples:")
    print("\n1. JSON format (products.json):")
    print('''[
  {
    "asin": "B0C5HYBQMW",
    "title": "BLUETTI Power Station",
    "affiliate_url": "https://amzn.to/4ruXr7n"
  },
  {
    "asin": "B0CL66FYLQ",
    "title": "BLUETTI Solar Generator",
    "affiliate_url": "https://amzn.to/3KeqtLE"
  }
]''')
    print("\n2. Plain text format (products.txt):")
    print('''# Lines starting with # are comments
B0C5HYBQMW|BLUETTI Power Station|https://amzn.to/4ruXr7n
B0CL66FYLQ|BLUETTI Solar Generator|https://amzn.to/3KeqtLE
B0BVLPGS79|EF ECOFLOW Power|https://amzn.to/47D9dI1

# Or just ASINs (one per line):
B0C5HYBQMW
B0CL66FYLQ''')
    print("\n3. Excel format (products.xlsx):")
    print("   Column A: asin")
    print("   Column B: title")
    print("   Column C: affiliate_url")
    exit(1)

# Process products (rest of the code remains the same)
results = []
print(f"\n🚀 Starting to process {len(products)} products...")

for i, product in enumerate(products, 1):
    asin = product['asin']
    print(f"\n📦 Processing {i}/{len(products)}: {asin}")
    
    try:
        # Call Amazon API
        response = amazon_api.get_items([asin])
        
        if not response:
            raise Exception("No response from Amazon API")
        
        # Check for errors in response
        if 'Errors' in response:
            error_msg = response['Errors'][0].get('Message', 'Unknown API error')
            raise Exception(f"API Error: {error_msg}")
        
        if 'ItemsResult' not in response or 'Items' not in response['ItemsResult']:
            raise Exception("Invalid response structure")
        
        items = response['ItemsResult']['Items']
        if not items:
            raise Exception("No items returned")
        
        item = items[0]
        
        # Extract price
        price = "N/A"
        try:
            offers = item.get('Offers', {})
            listings = offers.get('Listings', [])
            if listings:
                price_info = listings[0].get('Price', {})
                price = price_info.get('DisplayAmount', price_info.get('Amount', 'N/A'))
        except Exception as price_error:
            print(f"  ⚠️ Price extraction failed: {price_error}")
        
        # Extract title
        title = product.get('title', 'N/A')
        try:
            item_info = item.get('ItemInfo', {})
            title_info = item_info.get('Title', {})
            api_title = title_info.get('DisplayValue')
            if api_title:
                title = api_title
        except Exception as title_error:
            print(f"  ⚠️ Title extraction failed: {title_error}")
        
        # Create result - CHANGED affiliate_link to affiliate_url
        result = {
            "asin": asin,
            "title": title,
            "affiliate_url": product.get("affiliate_url", "N/A"),  # Changed from affiliate_link
            "price": price,
            "last_updated": datetime.datetime.utcnow().isoformat()
        }
        
        results.append(result)
        print(f"  ✅ Success: Price = {price}")
        
        # Rate limiting - Amazon allows 1 request per second for free tier
        time.sleep(1)
        
    except Exception as e:
        error_msg = str(e)
        print(f"  ❌ Failed: {error_msg}")
        
        result = {
            "asin": asin,
            "title": product.get("title", "N/A"),
            "affiliate_url": product.get("affiliate_url", "N/A"),  # Changed from affiliate_link
            "price": "N/A",
            "error": error_msg,
            "last_updated": datetime.datetime.utcnow().isoformat()
        }
        results.append(result)

# Save results
print(f"\n💾 Saving results to prices.json...")
try:
    with open("prices.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print("✅ File saved successfully")
except Exception as e:
    print(f"❌ Error saving file: {e}")

# Summary
successful = sum(1 for r in results if r.get('price') != 'N/A' and 'error' not in r)
failed = len(results) - successful

print(f"\n📊 SUMMARY:")
print(f"  Total processed: {len(results)}")
print(f"  Successful: {successful}")
print(f"  Failed: {failed}")

if failed > 0:
    print(f"\n❌ Failed items:")
    for r in results:
        if 'error' in r:
            print(f"  - {r['asin']}: {r['error']}")

print(f"\n🎉 Process completed!")
